import abc
import json
import logging
from typing import List, Dict, Any, Generator
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from datetime import datetime
import re

logger = logging.getLogger(__name__)

class BaseImporter(abc.ABC):
    """
    Abstract Base Class for all File Importers.
    """
    @abc.abstractmethod
    def read_headers(self, file_path: str) -> List[str]:
        """Read and return column headers from the file."""
        pass

    @abc.abstractmethod
    def import_data(self, file_path: str, mapping: Dict[str, str] = None) -> Generator[Dict[str, Any], None, None]:
        """
        Yield rows as dictionaries based on the provided mapping.
        mapping: { 'DbField': 'FileHeader' }
        """
        pass

class ExcelImporter(BaseImporter):
    def _serialize_value(self, val):
        """Convert values to JSON-serializable types."""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.isoformat()
        return val

    def read_headers(self, file_path: str) -> List[str]:
        if hasattr(file_path, 'seek'):
            file_path.seek(0)
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        # Read the first row
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip() for h in row if h is not None]
            break
        # wb.close() # Avoid closing externally provided file objects if possible, or just accept it.
        # Openpyxl read-only mode uses a file handle. If we close wb, it might close the handle.
        # But we need to close wb to release resources.
        wb.close()
        return headers

    def import_data(self, file_path: str, mapping: Dict[str, str] = None) -> Generator[Dict[str, Any], None, None]:
        """
        mapping: { 'db_field': 'Excel Header Name' }
        """
        if hasattr(file_path, 'seek'):
            file_path.seek(0)
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 1. Identify Column Indices for the Mapping
        # Get headers first
        file_headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            file_headers = [str(h).strip() for h in row if h is not None]
            break
        
        # Build index map: {'Excel Header': 0, ...}
        header_map = {name: i for i, name in enumerate(file_headers)}
        
        # Invert user mapping to get col indices: {'db_field': col_index}
        db_to_col_index = {}
        if mapping:
            for db_field, excel_header in mapping.items():
                if excel_header in header_map:
                    db_to_col_index[db_field] = header_map[excel_header]
                else:
                    logger.warning(f"Mapped header '{excel_header}' not found in file.")

        # 2. Iterate Data
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue  # Skip empty rows
            
            data = {}
            if not mapping:
                # If no mapping provided, yield raw dict using headers
                for i, val in enumerate(row):
                    if i < len(file_headers):
                        data[file_headers[i]] = self._serialize_value(val)
            else:
                # Use mapping to extract specific fields
                for db_field, col_idx in db_to_col_index.items():
                    if col_idx < len(row):
                        data[db_field] = self._serialize_value(row[col_idx])
            
            yield data
        
        wb.close()


class MSPImporter(BaseImporter):
    """
    Microsoft Project XML Parser.
    Extracts Tasks, UIDs, Dates, Progress.
    Does NOT need detailed mapping as XML schema is standard.
    """
    ns = {'msp': 'http://schemas.microsoft.com/project'} # Standard 2007+ schema usually

    def read_headers(self, file_path: str) -> List[str]:
        # Return standard MSP fields we support
        return ['UID', 'Name', 'Start', 'Finish', 'PercentComplete', 'WBS', 'OutlineNumber']

    def import_data(self, file_path: str, mapping: Dict[str, str] = None) -> Generator[Dict[str, Any], None, None]:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        # Handle namespaces if present (simple hack: strict namespace handling is pain in generic parsers)
        # We'll use local-name() or loop and strip tags if namespace is standard
        # But 'http://schemas.microsoft.com/project' is common.
        
        # Find 'Task' elements
        # Note: root usually has {ns}Tasks -> {ns}Task
        
        # Fallback recursive search for 'Task' tag
        tasks = []
        # Try finding all elements ending in 'Task'
        for elem in root.iter():
            if elem.tag.endswith('Task'):
                tasks.append(elem)
        
        for task in tasks:
            data = {}
            for child in task:
                tag = child.tag.split('}')[-1] # Strip NS
                if tag in ['UID', 'Name', 'Start', 'Finish', 'PercentComplete', 'WBS', 'OutlineNumber']:
                    data[tag] = child.text
            
            # Skip summary tasks if needed, or identifying them
            # Usually Summary tasks have children or specific flags. 
            # We'll yield everything, logic layer can filter.
            if 'Name' in data:
                 yield data

class P6Importer(BaseImporter):
    """
    Primavera P6 (.xer) Parser.
    Proprietary text format. We focus on %T=TASK table.
    """
    def read_headers(self, file_path: str) -> List[str]:
        # P6 headers depend on the table definition in the file.
        # We basically say we support standard P6 Task fields.
        return ['task_code', 'task_name', 'target_start_date', 'target_end_date', 'act_start_date', 'act_end_date', 'phys_complete_pct']

    def import_data(self, file_path: str, mapping: Dict[str, str] = None) -> Generator[Dict[str, Any], None, None]:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        current_table = None
        headers = []
        for line in lines:
            line = line.strip()
            parts = line.split('\t')
            row_type = parts[0]
            
            if row_type == '%T':
                current_table = parts[1]
                # We only care about TASK table
            elif row_type == '%F':
                if current_table == 'TASK':
                    headers = parts[1:] # Field names
            elif row_type == '%R':
                if current_table == 'TASK':
                    values = parts[1:]
                    row_data = dict(zip(headers, values))
                    yield row_data
